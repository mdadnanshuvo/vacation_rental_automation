import openpyxl
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
import os

EXCEL_FILE_PATH = "data/test_results.xlsx"


def ensure_data_folder_exists():
    """Ensure the data folder exists."""
    os.makedirs(os.path.dirname(EXCEL_FILE_PATH), exist_ok=True)


def create_excel_file():
    """Create a new Excel file with the specified columns."""
    ensure_data_folder_exists()  # Ensure the data folder exists

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Test Results"

    # Define the column headers
    columns = ["Key", "Url", "Page", "Test Case", "Passed", "Comments"]
    for col_num, column_title in enumerate(columns, 1):
        col_letter = get_column_letter(col_num)
        sheet[f'{col_letter}1'] = column_title

    workbook.save(EXCEL_FILE_PATH)
    print(f"Excel file '{EXCEL_FILE_PATH}' created successfully.")


def append_to_excel_file(data, additional_info):
    """
    Append a new row of data to the existing Excel file.

    Args:
        data (list): The main row data (Key, Url, Page, Test Case, Passed, Comments).
        additional_info (dict): A dictionary containing additional information like location and date range.

        Example of `additional_info`:
            {
                "location": "New York",
                "check_in_date": "2025-01-15",
                "check_out_date": "2025-01-20"
            }
    """
    if not os.path.exists(EXCEL_FILE_PATH):
        create_excel_file()  # Create the Excel file if it does not exist

    workbook = openpyxl.load_workbook(EXCEL_FILE_PATH)
    sheet = workbook.active

    # Extract additional information
    location = additional_info.get("location", "Unknown Location")
    check_in_date = additional_info.get(
        "check_in_date", "Unknown Check-in Date")
    check_out_date = additional_info.get(
        "check_out_date", "Unknown Check-out Date")

    # Format the additional details professionally
    formatted_comment = (
        f"{data[5]} | "
        f"**Location:** {location} | "
        f"**Date Range:** Check-in: {
            check_in_date}, Check-out: {check_out_date}"
    )
    data[5] = formatted_comment  # Update the comments in the data list

    # Append the updated data to the sheet
    sheet.append(data)
    workbook.save(EXCEL_FILE_PATH)
    print(f"Data appended to Excel file '{EXCEL_FILE_PATH}' successfully.")
